import json
import random
from django.contrib.auth import login, authenticate
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Avg
from datetime import timedelta
from DeliveryDjango import settings
from DeliveryDjango.forms import OrderForm, RegistrationForm, LoginForm, ReviewForm
from DeliveryDjango.models import Restaurant, Dish, MenuItem, Cart, CartItem, Order, Person, Review, OrderItem, Category
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Count, Sum


def restaurant_list(request):
    restaurants = Restaurant.objects.all()  # Получаем все объекты Restaurant
    return render(request, 'restaurant_list.html', {'restaurants': restaurants})


def menu_restaurant(request, restaurant_id):
    request.session['selected_restaurant'] = restaurant_id
    restaurant = Restaurant.objects.get(id=restaurant_id)
    dishes = Dish.objects.filter(restaurant=restaurant)

    # Группируем блюда по категориям и создаем словарь
    grouped_dishes_by_category = {}
    for dish in dishes:
        category_name = dish.category.name
        if category_name not in grouped_dishes_by_category:
            grouped_dishes_by_category[category_name] = []
        grouped_dishes_by_category[category_name].append(dish)

    context = {
        'grouped_dishes_by_category': grouped_dishes_by_category,
    }
    return render(request, 'restaurant_menu.html', context)


class DishDetail(generic.DetailView):
    # This line is the reason why we can access dish in dish_detail.html
    model = Dish
    template_name = 'dish_detail.html'


def add_to_cart(request, dish_id):
    # Получаем текущего пользователя
    user = request.user
    # Получаем блюдо по ID
    dish = Dish.objects.get(pk=dish_id)

    # Получаем корзину текущего пользователя, создав новую, если она еще не существует
    if user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=user)
    else:
        cart, created = Cart.objects.get_or_create(id=request.session.get('cart_id'))
        request.session['cart_id'] = cart.id

    # Добавляем блюдо в корзину
    cart.add_item(dish, 1)  # Можно изменить количество

    # Обновляем общую стоимость корзины
    cart.update_total_price()

    # Перенаправляем обратно на страницу продукта
    return redirect('view_cart')


def view_cart(request):
    # Получаем текущего пользователя
    user = request.user
    if user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=user)
        cart_items = cart.cartitem_set.all()
    else:
        # Получаем корзину из сессии
        cart_id = request.session.get('cart_id')
        if cart_id:
            cart = Cart.objects.get(id=cart_id)
            cart_items = cart.cartitem_set.all()
        else:
            # Отображаем пустую корзину
            cart_items = []
    total_price = sum(ci.item.price * ci.quantity for ci in cart_items)
    return render(request, 'cart.html', {'cart_items': cart_items, 'total_price': total_price})


def order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            user = request.user
            # Проверяем, является ли пользователь анонимным
            if user.is_anonymous:
                # Создаем анонимного пользователя для целей заказа
                anonymous_user = User.objects.get_or_create(username='AnonymousUser', password='AnonPassword',
                                                            is_active=False)
                # anonymous_user.save()
                user = anonymous_user[0]  # Здесь нужно указать [0] для того, чтоб получить именно пользователя
                # Потому что метод get_or_create возвращает список (User, True/False)
                # В зависимости от того, найден ли пользователь или нет

                cart = Cart.objects.get(id=request.session.get('cart_id'))
            else:
                # Если пользователь уже авторизован, используем его
                cart = Cart.objects.get(user=user)

            # Сохраняем данные формы в переменные
            name = form.cleaned_data['name']
            last_name = form.cleaned_data['last_name']
            phone_number = form.cleaned_data['phone_number']
            address = form.cleaned_data['address']
            payment_type = form.cleaned_data['payment_type']

            restaurant = Restaurant.objects.get(id=request.session.get('selected_restaurant'))

            # Создаем новый объект Order
            new_order = Order(
                user=user,  # Теперь user может быть экземпляром User или AnonymousUser
                payment_method=payment_type,
                total_price=cart.total_price,
                restaurant=restaurant,  # Используем ID ресторана из сессии
                created_at=None,  # Не устанавливаем дату создания, так как она автоматически заполняется
            )

            # Сохраняем объект Order
            new_order.save()

            # Получаем все товары из корзины
            cart_items = CartItem.objects.filter(cart=cart)

            # Добавление элементов корзины в заказ
            for cart_item in cart_items:
                OrderItem.objects.create(order=new_order, cart_item=cart_item)

            # Удаляем все блюда из корзины
            for item in cart.items.all():
                cart.remove_item(item)

            # Здесь можно добавить логику для обработки способа оплаты и других действий

            # Перенаправляем пользователя на страницу подтверждения заказа
            return redirect('thank_you')
    else:
        form = OrderForm()

    return render(request, 'order.html', {'form': form})


def thank_you(request):
    return render(request, 'thank_you.html')


def calculate_total_price_from_cart(request):
    user = request.user
    # Получаем все элементы корзины для текущего пользователя
    if user.is_anonymous:
        cart_items = CartItem.objects.filter(id=request.session.get('cart_id'))
    else:
        cart_items = CartItem.objects.filter(user=user)

    # Инициализируем общую сумму заказа
    total_price = 0

    # Проходимся по всем элементам корзины
    for item in cart_items:
        # Умножаем цену товара на его количество и добавляем к общей сумме
        total_price += item.item.price * item.quantity

    return total_price


def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['email'],  # используем email в качестве имени пользователя
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password1']
            )
            person = Person.objects.create(
                user=user,
                name=form.cleaned_data['name'],
                lastname=form.cleaned_data['lastname'],
                middlename=form.cleaned_data['middlename'],
                phone=form.cleaned_data['phone'],
                email=form.cleaned_data['email'],
                pin=generate_unique_pin()
            )
            return redirect('restaurant_list')
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})


def login_view(request):
    form = LoginForm(request.POST or None)
    if form.is_valid():
        email = form.cleaned_data.get('email')
        password = form.cleaned_data.get('password')
        user = settings.authenticate(email=email, password=password)
        if user is not None:
            login(request, user)
            return redirect('restaurant_list')
    return render(request, 'login.html', {'form': form})


def generate_unique_pin():
    """Генерирует уникальный PIN из четырех случайных цифр."""
    while True:
        pin = ''.join([str(random.randint(0, 9)) for _ in range(4)])
        if not Person.objects.filter(pin=pin).exists():
            return pin


def user_orders(request):
    if request.user.is_authenticated:
        orders = Order.objects.filter(user=request.user)
        reviews = Review.objects.filter(order__in=orders)
        return render(request, 'user_orders.html', {'orders': orders, 'reviews': reviews})
    else:
        return redirect('login')  # Перенаправление на страницу входа, если пользователь не авторизован


def rate_order(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review, created = Review.objects.get_or_create(
                order=order,
                defaults={
                    'user': request.user,
                    'review_text': form.cleaned_data['review_text'],
                    'delivery_speed_rating': form.cleaned_data['delivery_speed_rating'],
                    'taste_intensity_rating': form.cleaned_data['taste_intensity_rating'],
                    'product_quality_rating': form.cleaned_data['product_quality_rating'],
                    'service_quality_rating': form.cleaned_data['service_quality_rating']
                }
            )
            order.have_review = True
            order.save()
            if not created:
                review.review_text = form.cleaned_data['review_text']
                review.delivery_speed_rating = form.cleaned_data['delivery_speed_rating']
                review.taste_intensity_rating = form.cleaned_data['taste_intensity_rating']
                review.product_quality_rating = form.cleaned_data['product_quality_rating']
                review.service_quality_rating = form.cleaned_data['service_quality_rating']
                review.save()
            return redirect('user_orders')
    else:
        form = ReviewForm()
    return render(request, 'rate_order.html', {'form': form, 'order': order})


@csrf_exempt
def orders_chart_data(request):
    orders_with_reviews = Order.objects.filter(have_review=True).count()
    orders_without_reviews = Order.objects.filter(have_review=False).count()
    return JsonResponse({
        'orders_with_reviews': orders_with_reviews,
        'orders_without_reviews': orders_without_reviews
    })


def restaurant_list_view(request):
    # Логика для получения запроса и значений фильтра
    selected_filter_value = request.GET.get('restaurant', 'all')

    # Получение всех объектов Restaurant
    restaurants = Restaurant.objects.all()

    # Создание контекста с данными для шаблона
    context = {
        'cl': {'opts': {'app_label': 'deliverydjango', 'model_name': 'review'}},
        # Обновите 'your_app' и 'your_model' соответственно
        'selected_filter_value': selected_filter_value,
        'restaurants': restaurants,
    }
    return render(request, 'restaurant_list.html', context)


@csrf_exempt
def review_ratings_chart_data(request, restaurant_id):
    now = timezone.now()

    def get_ratings_for_period(start_time):
        reviews = Review.objects.filter(order__restaurant_id=restaurant_id, order__created_at__gte=start_time)
        ratings = reviews.aggregate(
            avg_delivery_speed=Avg('delivery_speed_rating'),
            avg_taste_intensity=Avg('taste_intensity_rating'),
            avg_product_quality=Avg('product_quality_rating'),
            avg_service_quality=Avg('service_quality_rating'),
        )
        return {
            'avg_delivery_speed': ratings['avg_delivery_speed'] or 0,
            'avg_taste_intensity': ratings['avg_taste_intensity'] or 0,
            'avg_product_quality': ratings['avg_product_quality'] or 0,
            'avg_service_quality': ratings['service_quality'] or 0,
        }

    data = {
        'day': get_ratings_for_period(now - timedelta(days=1)),
        'week': get_ratings_for_period(now - timedelta(weeks=1)),
        'month': get_ratings_for_period(now - timedelta(days=30)),
    }

    return JsonResponse(data)


@csrf_exempt
def restaurants_rating_data(request):
    ratings = []
    for order in Order.objects.all():
        reviews = Review.objects.filter(order=order)
        total_rating = sum(
            review.delivery_speed_rating + review.taste_intensity_rating + review.product_quality_rating + review.service_quality_rating for review in
            reviews)
        ratings.append({
            'name': order.restaurant.name,
            'rating': total_rating / len(reviews) if reviews else 0
        })
    return JsonResponse(ratings, safe=False)


@csrf_exempt
def restaurant_review_stats(request, restaurant_id):
    restaurant = Restaurant.objects.get(pk=restaurant_id)
    reviews = Review.objects.filter(order__restaurant=restaurant)

    # Calculate integer rating averages
    delivery_speed_avg = reviews.aggregate(Avg('delivery_speed_rating'))['delivery_speed_rating__avg']
    taste_intensity_avg = reviews.aggregate(Avg('taste_intensity_rating'))['taste_intensity_rating__avg']
    product_quality_avg = reviews.aggregate(Avg('product_quality_rating'))['product_quality_rating__avg']
    service_quality_avg = reviews.aggregate(Avg('service_quality_rating'))['service_quality_rating__avg']

    # Prepare data in dictionary format for Chart.js
    chart_data = {
        'labels': ['Скорость доставки', 'Интенсивность вкуса', 'Качество продукции', 'Качество обслуживания'],
        'datasets': [{
            'label': 'Средний рейтинг',
            'data': [delivery_speed_avg or 0, taste_intensity_avg or 0, product_quality_avg or 0, service_quality_avg or 0],
            'backgroundColor': ['#FFCE56', '#36A2EB', '#FF7F0E', '#00FFFF'],
            'borderColor': ['#FFCE56', '#36A2EB', '#FF7F0E', '#00FFFF']
        }]
    }

    # Convert dictionary to JSON format
    chart_data_json = json.dumps(chart_data)

    return JsonResponse({'chart_data': chart_data_json})


@csrf_exempt
def all_restaurants_review_stats(request):
    reviews = Review.objects.all()

    reviews_by_restaurant = reviews.values('order__restaurant__name').annotate(
        delivery_speed_avg=Avg('delivery_speed_rating'),
        taste_intensity_avg=Avg('taste_intensity_rating'),
        product_quality_avg=Avg('product_quality_rating'),
        service_quality_avg=Avg('service_quality_rating'),
    )

    chart_data = {
        'labels': [restaurant['order__restaurant__name'] for restaurant in reviews_by_restaurant],
        'datasets': [{
            'label': 'Средний рейтинг',
            'data': [(restaurant['delivery_speed_avg'] +
                        restaurant['taste_intensity_avg'] +
                        restaurant['product_quality_avg'] +
                        restaurant['service_quality_avg']) / 3 or 0 for restaurant in reviews_by_restaurant],
            'backgroundColor': 'rgba(0, 123, 255, 0.2)',
            'borderColor': 'rgba(0, 123, 255, 1)',
            'borderWidth': 1,
            'fill': False,
        }]
    }

    return JsonResponse(chart_data)

def export_reviews_to_excel(request):
    # Создаем новый workbook и активный лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reviews'

    # Устанавливаем заголовки столбцов
    columns = ['Ресторан', 'Пользователь', 'Отзыв', 'Рейтинг времени доставки', 'Рейтинг вкуса', 'Рейтинг качества продукта', 'Рейтинг качества обслуживания']
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        ws[column_letter + '1'] = column_title

    # Заполняем данные
    reviews = Review.objects.all()
    for row_num, review in enumerate(reviews, 2):
        try:
            restaurant_name = review.order.restaurant.name if review.order.restaurant else 'No Restaurant'
        except Restaurant.DoesNotExist:
            restaurant_name = 'No Restaurant'
        ws['A' + str(row_num)] = restaurant_name
        ws['B' + str(row_num)] = review.user.username
        ws['C' + str(row_num)] = review.review_text
        ws['D' + str(row_num)] = review.delivery_speed_rating
        ws['E' + str(row_num)] = review.taste_intensity_rating
        ws['F' + str(row_num)] = review.product_quality_rating
        ws['G' + str(row_num)] = review.service_quality_rating

    # Устанавливаем правильные размеры столбцов
    for col_num in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].auto_size = True

    # Создаем HTTP ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reviews.xlsx'
    wb.save(response)
    return response

def export_reviews_view(request):
    return render(request, 'admin/export_reviews.html')
@csrf_exempt
def restaurants_with_reviews_stats(request):
    with_reviews = Restaurant.objects.filter(order__have_review=True).distinct().count()
    without_reviews = Restaurant.objects.filter(order__have_review=False).distinct().count()
    rest = Restaurant.objects.filter(order__have_review=True).all()
    
    
    return JsonResponse({'with_reviews': with_reviews, 'without_reviews': without_reviews})
@csrf_exempt
def orders_by_restaurant_stats(request):
    orders_by_restaurant = Order.objects.values('restaurant__name').annotate(order_count=Count('id'))
    data = {
        'labels': [entry['restaurant__name'] for entry in orders_by_restaurant],
        'datasets': [{
            'label': 'Число заказов',
            'data': [entry['order_count'] for entry in orders_by_restaurant],
            'backgroundColor': 'rgba(54, 162, 235, 0.2)',
            'borderColor': 'rgba(54, 162, 235, 1)',
            'borderWidth': 1
        }]
    }
    return JsonResponse(data)
from django.http import JsonResponse

@csrf_exempt
def review_counts_by_restaurant(request, restaurant_id):
    orders_with_reviews = Order.objects.filter(restaurant_id=restaurant_id, have_review=True).count()
    orders_without_reviews = Order.objects.filter(restaurant_id=restaurant_id, have_review=False).count()
    data = {
        'with_reviews': orders_with_reviews,
        'without_reviews': orders_without_reviews
    }
    return JsonResponse(data)
@csrf_exempt
def top_restaurants_by_reviews(request):
    top_restaurants = (
        Restaurant.objects.annotate(
            total_reviews=Count('order__reviews'),
            sum_rating=Sum('order__reviews__delivery_speed_rating') +
                       Sum('order__reviews__taste_intensity_rating') +
                       Sum('order__reviews__product_quality_rating') +
                       Sum('order__reviews__service_quality_rating')
        )
        .order_by('-total_reviews', '-sum_rating')[:3]
        .values('name', 'total_reviews', 'sum_rating')
    )
    return JsonResponse(list(top_restaurants), safe=False)
